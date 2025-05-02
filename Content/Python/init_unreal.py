import unreal
import pypinyin
import pandas as pd
import json
import io
import os
import openpyxl
from openpyxl import load_workbook
import traceback

@unreal.uclass()
class PythonBridgeImplementation(unreal.PythonBridge):

    @unreal.ufunction(override=True)
    def function_implemented_in_python(self):
        unreal.log_warning("Wow! 111")

    @unreal.ufunction(override=True)
    def test(self):
        unreal.log_warning("Wow! Test!!")
        return "Test"

    @unreal.ufunction(override=True)
    def to_pinyin(self, in_text):
        # 将中文字符串转换为拼音
        pinyin_text = pypinyin.lazy_pinyin(in_text)
        # 将拼音列表转换为字符串, 并且每个拼音的首字母大写
        pinyin_string = ''.join([s[0].upper() + s[1:] for s in pinyin_text])
        unreal.log_warning(pinyin_string)
        return pinyin_string
        
    @unreal.ufunction(override=True)
    def excel_to_csv(self, excel_path):
        try:
            # 读取Excel文件
            df = pd.read_excel(excel_path)
            
            # 将DataFrame转换为CSV字符串
            csv_string = df.to_csv(index=False)
            
            unreal.log_warning(f"成功从Excel转换为CSV，行数: {len(df)}")
            return csv_string
        except Exception as e:
            error_msg = f"Excel转换CSV出错: {str(e)}"
            unreal.log_error(error_msg)
            return ""
            
    @unreal.ufunction(override=True)
    def excel_to_json(self, excel_path, sheet_name):
        try:
            # 读取Excel文件
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # 将DataFrame转换为JSON字符串
            json_string = df.to_json(orient="records")
            
            unreal.log_warning(f"成功从Excel转换为JSON，行数: {len(df)}")
            return json_string
        except Exception as e:
            error_msg = f"Excel转换JSON出错: {str(e)}"
            unreal.log_error(error_msg)
            return ""
            
    @unreal.ufunction(override=True)
    def csv_to_excel(self, csv_string, excel_path, sheet_name):
        try:
            # 从CSV字符串创建DataFrame
            df = pd.read_csv(io.StringIO(csv_string), encoding='utf-8')
            
            # 检查Excel文件是否已存在
            file_exists = os.path.isfile(excel_path)
            
            if file_exists:
                try:
                    # 使用更安全的方式处理Excel文件
                    book = load_workbook(excel_path)
                    
                    # 如果sheet已存在，则删除它
                    if sheet_name in book.sheetnames:
                        idx = book.sheetnames.index(sheet_name)
                        book.remove(book[sheet_name])
                        # 确保删除后保存工作簿
                        book.save(excel_path)
                    
                    # 重新加载工作簿，使用openpyxl引擎以支持Unicode
                    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 格式化Excel文件，设置表头和类型行格式
                    try:
                        self._format_excel_sheet(excel_path, sheet_name)
                    except Exception as e:
                        unreal.log_warning(f"设置Excel格式失败: {str(e)}")
                    
                    unreal.log_warning(f"成功将CSV写入到Excel文件的工作表 {sheet_name}，保存到: {excel_path}")
                    return True
                    
                except Exception as e:
                    # 记录完整的错误堆栈
                    error_stack = traceback.format_exc()
                    unreal.log_error(f"处理Excel文件出错: {str(e)}\n{error_stack}")
                    
                    # 备份方案：如果上面的方法失败，尝试创建新文件
                    try:
                        # 如果文件存在且可以读取，则尝试保留其他Sheet
                        if os.path.isfile(excel_path):
                            # 读取所有其他工作表
                            xls = pd.ExcelFile(excel_path)
                            all_sheets = xls.sheet_names
                            
                            # 创建新的ExcelWriter，使用openpyxl引擎
                            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                                # 写入当前DataFrame
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                                
                                # 保留其他工作表
                                for other_sheet in all_sheets:
                                    if other_sheet != sheet_name:
                                        other_df = pd.read_excel(excel_path, sheet_name=other_sheet)
                                        other_df.to_excel(writer, sheet_name=other_sheet, index=False)
                            
                            # 尝试格式化Excel
                            try:
                                self._format_excel_sheet(excel_path, sheet_name)
                            except Exception as e:
                                unreal.log_warning(f"设置Excel格式失败: {str(e)}")
                                
                            unreal.log_warning(f"使用备用方法将CSV写入到Excel，保存到: {excel_path}")
                            return True
                        else:
                            # 文件不存在时直接创建
                            df.to_excel(excel_path, sheet_name=sheet_name, index=False)
                            
                            # 尝试格式化Excel
                            try:
                                self._format_excel_sheet(excel_path, sheet_name)
                            except Exception as e:
                                unreal.log_warning(f"设置Excel格式失败: {str(e)}")
                                
                            unreal.log_warning(f"创建新Excel文件，工作表: {sheet_name}，保存到: {excel_path}")
                            return True
                    except Exception as e2:
                        # 最后的备份方案：直接覆盖创建
                        unreal.log_error(f"备用方法也失败，将直接创建新文件: {str(e2)}")
                        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
                        return True
            else:
                # 文件不存在，直接创建新文件，使用openpyxl引擎
                df.to_excel(excel_path, sheet_name=sheet_name, index=False, engine='openpyxl')
                
                # 尝试格式化Excel
                try:
                    self._format_excel_sheet(excel_path, sheet_name)
                except Exception as e:
                    unreal.log_warning(f"设置Excel格式失败: {str(e)}")
                    
                unreal.log_warning(f"创建新Excel文件，工作表: {sheet_name}，保存到: {excel_path}")
                return True
                
        except Exception as e:
            error_msg = f"CSV转换Excel出错: {str(e)}"
            unreal.log_error(error_msg)
            return False
            
    def _format_excel_sheet(self, excel_path, sheet_name):
        """格式化Excel工作表，美化表头和类型行的外观"""
        try:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                
                # 获取最大列数
                max_col = sheet.max_column
                
                # 设置表头样式（第一行）
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col)
                    # 如果是第一列且值为"Row_Name"，特殊处理
                    if col == 1 and cell.value == "Row_Name":
                        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")  # 白色加粗
                    else:
                        cell.font = openpyxl.styles.Font(bold=True)  # 加粗
                    cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 蓝色背景
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")  # 居中对齐
                
                # 设置类型行样式（第二行）
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=2, column=col)
                    cell.font = openpyxl.styles.Font(italic=True, color="666666")  # 灰色斜体
                    cell.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # 浅灰色背景
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")  # 居中对齐
                
                # 自动调整列宽
                for col in range(1, max_col + 1):
                    # 获取该列所有单元格的值长度
                    max_length = 0
                    for row in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    # 设置列宽（最大长度+2，确保有一些空间）
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
                
                # 冻结表头和类型行
                sheet.freeze_panes = "A3"
                
                # 保存修改
                book.save(excel_path)
                
                # 添加成功日志
                unreal.log_warning(f"成功格式化Excel工作表: {sheet_name}")
                
        except Exception as e:
            # 记录错误但不中断流程
            unreal.log_warning(f"格式化Excel出错: {str(e)}")
            
    @unreal.ufunction(override=True)
    def json_to_excel(self, json_string, excel_path, sheet_name):
        try:
            # 解析JSON字符串
            data = json.loads(json_string)
            
            # 从JSON创建DataFrame
            df = pd.DataFrame(data)
            
            # 复用csv_to_excel的逻辑，将DataFrame转为CSV再处理
            csv_string = df.to_csv(index=False)
            return self.csv_to_excel(csv_string, excel_path, sheet_name)
                
        except Exception as e:
            error_msg = f"JSON转换Excel出错: {str(e)}"
            unreal.log_error(error_msg)
            return False